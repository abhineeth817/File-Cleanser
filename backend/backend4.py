
import os
import threading
import time
import docx
import csv
import backend34 as b34
from flask import app, jsonify, request, send_file, Flask
from flask_cors import CORS
import openpyxl
from presidio_analyzer import AnalyzerEngine
from presidio_anonymizer import AnonymizerEngine
from streamlit import progress
import granite_extracion as ge

# Presidio Analyzer and Anonymizer
analyzer = AnalyzerEngine()
anonymizer = AnonymizerEngine()

# Flask app setup
app = Flask(__name__)
CORS(app)

# Simple in-memory progress tracker
progress = {}


# --- Cleanse / Anonymize text using Presidio ---
def cleanse_text(text, *args, **kwargs):
    """Anonymizes PII from the given text using Presidio."""
    if not text or text.isspace():
        return ""
    # Remove IP addresses using regex
    import re
    ipv4_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
    ipv6_pattern = r'\b(?:[A-Fa-f0-9]{1,4}:){7}[A-Fa-f0-9]{1,4}\b'
    text = re.sub(ipv4_pattern, '<IP_ADDRESS>', text)
    text = re.sub(ipv6_pattern, '<IP_ADDRESS>', text)
    results = analyzer.analyze(text=text, language="en")
    anonymized_result = anonymizer.anonymize(text=text, analyzer_results=results)
    return anonymized_result.text



def extract_docx(file_path):
    doc = docx.Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def extract_csv(file_path):
    text = ""
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text

def extract_txt(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()

def extract_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text

# Dispatcher
def process_file(file_path, upload_id=None, preserve_type=False):
    try:
        ext = os.path.splitext(file_path)[-1].lower()
        if ext in [".pdf", ".pptx", ".jpg", ".jpeg", ".png"]:
            raw = ge.extract_with_granite_docling(file_path, upload_id)
        
        elif ext in [".doc", ".docx"]:
            raw = extract_docx(file_path)
        elif ext == ".xlsx":
            raw = extract_xlsx(file_path)
        elif ext == ".csv":
            raw = extract_csv(file_path)
        elif ext == ".txt":
            raw = extract_txt(file_path)
        else:
            raw = "Unsupported file format."
        if upload_id:
            progress[upload_id] = 50
        if not (preserve_type and ext in [".csv", ".xlsx", ".pdf"]):
            result = cleanse_text(raw, upload_id)
        else:
            result = raw
        if upload_id:
            progress[upload_id] = 100
        return result
    except Exception as e:
        if upload_id:
            progress[upload_id] = -1
        raise e


# --- Flask API Endpoints ---

# --- File Upload Endpoint ---
@app.route("/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400
    # Save file temporarily
    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'uploads'))
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)
    temp_path = os.path.join(uploads_dir, file.filename)
    file.save(temp_path)
    upload_id = str(time.time())
    progress[upload_id] = 0
    preserve_type = request.form.get("preserve_type") == "true"
    ext = os.path.splitext(file.filename)[-1].lower()
    # Decide output path and format
    if preserve_type and ext in [".csv", ".xlsx", ".pdf"]:
        output_ext = ext
        output_path = os.path.join(uploads_dir, f"cleansed_{file.filename.replace('.', '_')}{output_ext}")
    else:
        output_ext = ".txt"
        output_path = os.path.join(uploads_dir, f"cleansed_{file.filename.replace('.', '_')}.txt")

    def worker():
        import logging
        start_time = time.time()
        timeout_seconds = 60 * 5  # 5 minutes
        try:
            app.logger.info(f"Worker started for upload_id={upload_id}, file={temp_path}")
            result = None
            try:
                result = process_file(temp_path, upload_id, preserve_type=preserve_type)
            except Exception as e:
                app.logger.error(f"process_file error for {upload_id}: {e}", exc_info=True)
                progress[upload_id] = -1
                return
            if result is None:
                app.logger.error(f"process_file returned None for {upload_id}")
                progress[upload_id] = -1
                return
            try:
                # Save cleansed file in requested format
                if preserve_type and ext == ".csv":
                    # Write CSV file
                    with open(output_path, "w", encoding="utf-8", newline='') as f:
                        for line in result.splitlines():
                            f.write(line + "\n")
                elif preserve_type and ext == ".xlsx":
                    # Cleanse each row as a single string for better context, then split back into cells
                    import openpyxl
                    wb_orig = openpyxl.load_workbook(temp_path)
                    wb_new = openpyxl.Workbook()
                    wb_new.remove(wb_new.active)
                    for sheet_name in wb_orig.sheetnames:
                        ws_orig = wb_orig[sheet_name]
                        ws_new = wb_new.create_sheet(title=sheet_name)
                        for row in ws_orig.iter_rows(values_only=True):
                            row_str = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                            cleansed_row_str = cleanse_text(row_str)
                            cleansed_row = cleansed_row_str.split('\t')
                            ws_new.append(cleansed_row)
                    wb_new.save(output_path)
                elif preserve_type and ext == ".pdf":
                    # Save as PDF using backend34's save_pdf
                    b34.save_pdf(result, output_path)
                else:
                    # Default: write as text
                    with open(output_path, "w", encoding="utf-8") as f:
                        f.write(result)
                app.logger.info(f"Output file written: {output_path}")
            except Exception as e:
                app.logger.error(f"Failed to write output file for {upload_id}: {e}", exc_info=True)
                progress[upload_id] = -1
                return
            if not os.path.exists(output_path):
                app.logger.error(f"Output file not found after writing for {upload_id}: {output_path}")
                progress[upload_id] = -1
                return
            progress[upload_id] = 100
            app.logger.info(f"Worker finished for upload_id={upload_id}")
        except Exception as e:
            app.logger.error(f"Error processing {upload_id}: {e}", exc_info=True)
            progress[upload_id] = -1
        finally:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                    app.logger.info(f"Temp file removed: {temp_path}")
                except Exception as e:
                    app.logger.error(f"Failed to remove temp file {temp_path}: {e}", exc_info=True)
            # Timeout check
            elapsed = time.time() - start_time
            if elapsed > timeout_seconds:
                app.logger.error(f"Worker timeout for upload_id={upload_id}")
                progress[upload_id] = -1
    threading.Thread(target=worker).start()
    # Return download URL with correct extension
    return jsonify({"upload_id": upload_id, "download_url": f"/download/{file.filename.replace('.', '_')}{output_ext}"})

# --- Progress Endpoint ---
@app.route("/progress/<upload_id>", methods=["GET"])
def get_progress(upload_id):
    value = progress.get(upload_id, 0)
    return jsonify({"progress": value})

# --- Download Endpoint ---
@app.route("/download/<filename>", methods=["GET"])
def download_result(filename):
    uploads_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'uploads'))
    # Support dynamic extension
    # Find file with .txt, .csv, .xlsx, or .pdf extension
    base_path = os.path.join(uploads_dir, f"cleansed_{filename}")
    for ext in [".txt", ".csv", ".xlsx", ".pdf"]:
        output_path = base_path + ext if not filename.endswith(ext) else base_path
        if os.path.exists(output_path):
            # Set correct mimetype
            if ext == ".csv":
                mimetype = "text/csv"
            elif ext == ".xlsx":
                mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif ext == ".pdf":
                mimetype = "application/pdf"
            else:
                mimetype = "text/plain"
            return send_file(output_path, as_attachment=True, download_name=f"cleansed_{filename}{ext}", mimetype=mimetype)
    return jsonify({"error": "File not ready"}), 404

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
